from openpyxl.reader.excel import load_workbook
import feedparser
import requests
import json
from bs4 import BeautifulSoup

wb = load_workbook('db.xlsx')
ws = wb.active

url = 'http://arunai.herokuapp.com/api/techfeeds' 
sourceLinksRSS = []
titles = []
pageLinks = []
images = []
newTitles = []
newPageLinks = []
newImages = []

for i, x in enumerate(ws['A']):
                titles.append(ws['A'+str(i+1)].value)
                pageLinks.append(ws['B'+str(i+1)].value)
                images.append(ws['C'+str(i+1)].value)
oldCount = len(titles)

try:
        for link in sourceLinksRSS:
                print('\nFecthing RSS from',link,'...')
                x = feedparser.parse(link)
                for element in x['items']:
                        try:
                            soup = BeautifulSoup(element['description'],'html.parser')
                            title = element['title']
                            pageLink = element['link']
                            image = soup.find('img')['src']
                            if(title not in titles):
                                titles.append(title)
                                pageLinks.append(pageLink)
                                images.append(image)
                        except:
                            print('Exception Handled')                           
                            
except requests.ConnectionError:
        print('No Internet Connection')

# To display & store the parsed titles
finally:
        newCount = len(titles)
        print('\n')
        for x in range(oldCount,newCount):
            ws['A'+str(x+1)].value = titles[x]
            ws['B'+str(x+1)].value = pageLinks[x]
            ws['C'+str(x+1)].value = images[x]
            newTitles.append(titles[x])
            newPageLinks.append(pageLinks[x])
            newImages.append(images[x])
            rp = requests.post(url, {
                "title": titles[x],
                "link": pageLinks[x],
                "image": images[x]
            })
            x+=1
            print(x, '. ', titles[x-1], pageLinks[x-1], images[x-1])
            
        wb.save('db.xlsx')
